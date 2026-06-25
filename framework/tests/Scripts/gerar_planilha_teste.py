"""
Gera uma planilha XLSX de teste para importação de contatos.
Colunas: NOME | CPF_CNPJ | TELEFONE | STATUS

Busca dados REAIS do banco hydraRemake para criar cenários de teste realistas:
- Contatos que JÁ EXISTEM no banco (para testar atualização)
- Contatos NOVOS fictícios (para testar criação)
- Telefones que JÁ EXISTEM (para testar atualização de status)
- Linhas com erros propositais (para testar validação)

Uso:
    pip install openpyxl mysql-connector-python
    python gerar_planilha_teste.py
"""

import random
import os
import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Conexão com o banco ──────────────────────────────────────────────────────
DB_CONFIG = {
    'host': '127.0.0.1',
    'database': 'hydraRemake',
    'user': 'root',
    'password': '',
    'charset': 'utf8mb4'
}

# ── Dados fictícios para contatos NOVOS ──────────────────────────────────────
NOMES_NOVOS_PF = [
    "Helena Nascimento", "Thiago Barros", "Isabela Pinto",
    "Diego Mendes", "Valentina Correia", "Enzo Cardoso",
    "Mariana Teixeira", "Leonardo Monteiro", "Sofia Ramos"
]

NOMES_NOVOS_PJ = [
    "Nova Distribuidora Norte LTDA", "Comércio Digital Express ME",
    "Soluções Integradas Tech EIRELI", "Atacadão Central Sul SA",
    "Oficina Mecânica Premium ME"
]

DDDS = ["11", "21", "31", "41", "47", "48", "51", "61", "71", "85"]


def gerar_cpf():
    nums = [random.randint(0, 9) for _ in range(9)]
    for _ in range(2):
        val = sum([(len(nums) + 1 - i) * v for i, v in enumerate(nums)]) % 11
        nums.append(0 if val < 2 else 11 - val)
    return ''.join(map(str, nums))


def gerar_cnpj():
    nums = [random.randint(0, 9) for _ in range(8)] + [0, 0, 0, 1]
    pesos1 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    val = sum(n * p for n, p in zip(nums, pesos1)) % 11
    nums.append(0 if val < 2 else 11 - val)
    pesos2 = [6] + pesos1
    val = sum(n * p for n, p in zip(nums, pesos2)) % 11
    nums.append(0 if val < 2 else 11 - val)
    return ''.join(map(str, nums))


def gerar_telefone():
    ddd = random.choice(DDDS)
    numero = f"9{''.join([str(random.randint(0,9)) for _ in range(8)])}"
    return ddd + numero


def buscar_dados_reais():
    """Busca contatos e telefones que já existem no banco de dados."""
    contatos_existentes = []
    telefones_existentes = []

    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True)

        # Busca contatos existentes (clientes + representantes)
        cursor.execute("""
            SELECT ce.NOME_CONTATO, ce.NUMERO_DOCUMENTO
            FROM CONTATO_EXTERNO ce
            WHERE ce.NOME_CONTATO IS NOT NULL AND ce.NOME_CONTATO != ''
            ORDER BY RAND()
            LIMIT 10
        """)
        contatos_existentes = cursor.fetchall()

        # Busca telefones existentes com seus vínculos e status
        cursor.execute("""
            SELECT t.NUM_TEL, t.CONFIRMADO, ce.NOME_CONTATO, ce.NUMERO_DOCUMENTO
            FROM TEL t
            JOIN CONTATO_TEL ct ON ct.ID_TEL = t.ID_TEL
            JOIN CONTATO_EXTERNO ce ON ce.ID_CONTATO_BLING = ct.ID_CONTATO_BLING
            ORDER BY RAND()
            LIMIT 8
        """)
        telefones_existentes = cursor.fetchall()

        cursor.close()
        conn.close()
        print(f"📊 Encontrados {len(contatos_existentes)} contatos e {len(telefones_existentes)} telefones no banco.")
    except mysql.connector.Error as err:
        print(f"⚠️  Erro ao conectar ao banco: {err}")
        print("   Gerando planilha apenas com dados fictícios...")

    return contatos_existentes, telefones_existentes


def gerar_linhas():
    """Gera linhas de teste cobrindo todos os cenários."""
    contatos_reais, telefones_reais = buscar_dados_reais()
    linhas = []

    # ─── GRUPO 1: Contatos que JÁ EXISTEM + telefone NOVO ────────────────────
    # Cenário: O sistema deve encontrar o contato e apenas criar/vincular o telefone novo
    for c in contatos_reais[:4]:
        linhas.append({
            'nome': c['NOME_CONTATO'],
            'cpf_cnpj': c['NUMERO_DOCUMENTO'] or '',
            'telefone': gerar_telefone(),
            'status': random.choice(["Confirmado", "Tentativa"]),
            'comentario': 'CONTATO EXISTENTE + TEL NOVO'
        })

    # ─── GRUPO 2: Telefones que JÁ EXISTEM → testar atualização de status ────
    # Cenário: Telefone existe com status X, planilha tem status Y → deve mostrar "Atualizar status"
    for t in telefones_reais[:4]:
        status_atual = "Confirmado" if t['CONFIRMADO'] else "Tentativa"
        # Inverte o status para forçar atualização
        novo_status = "Tentativa" if status_atual == "Confirmado" else "Confirmado"
        linhas.append({
            'nome': t['NOME_CONTATO'],
            'cpf_cnpj': t['NUMERO_DOCUMENTO'] or '',
            'telefone': t['NUM_TEL'],
            'status': novo_status,
            'comentario': f'TEL EXISTENTE ({status_atual}) → ATUALIZAR PARA {novo_status}'
        })

    # ─── GRUPO 3: Telefones que JÁ EXISTEM com MESMO status ──────────────────
    # Cenário: Nada muda, sistema deve avisar "Telefone já vinculado com mesmo status"
    for t in telefones_reais[4:6]:
        status_atual = "Confirmado" if t['CONFIRMADO'] else "Tentativa"
        linhas.append({
            'nome': t['NOME_CONTATO'],
            'cpf_cnpj': t['NUMERO_DOCUMENTO'] or '',
            'telefone': t['NUM_TEL'],
            'status': status_atual,
            'comentario': 'TEL EXISTENTE COM MESMO STATUS (sem mudança)'
        })

    # ─── GRUPO 4: Contatos NOVOS (não existem no banco) ──────────────────────
    # Cenário: Sistema deve criar contato novo + telefone novo
    for nome in NOMES_NOVOS_PF[:4]:
        linhas.append({
            'nome': nome,
            'cpf_cnpj': gerar_cpf(),
            'telefone': gerar_telefone(),
            'status': random.choice(["Confirmado", "Tentativa"]),
            'comentario': 'CONTATO NOVO (PF)'
        })

    for nome in NOMES_NOVOS_PJ[:2]:
        linhas.append({
            'nome': nome,
            'cpf_cnpj': gerar_cnpj(),
            'telefone': gerar_telefone(),
            'status': "Confirmado",
            'comentario': 'CONTATO NOVO (PJ)'
        })

    # ─── GRUPO 5: Contato existente SEM documento, busca por nome ─────────────
    if contatos_reais:
        c = contatos_reais[-1]
        linhas.append({
            'nome': c['NOME_CONTATO'],
            'cpf_cnpj': '',
            'telefone': gerar_telefone(),
            'status': 'Tentativa',
            'comentario': 'BUSCA POR NOME (sem doc)'
        })

    # ─── GRUPO 6: Contato novo SEM telefone ──────────────────────────────────
    # Cenário: Deve criar o contato mas não vincular telefone
    linhas.append({
        'nome': 'Empresa Sem Telefone LTDA',
        'cpf_cnpj': gerar_cnpj(),
        'telefone': '',
        'status': '',
        'comentario': 'SEM TELEFONE'
    })

    # ─── GRUPO 7: Linha com ERRO (nome e doc vazios) ─────────────────────────
    linhas.append({
        'nome': '',
        'cpf_cnpj': '',
        'telefone': gerar_telefone(),
        'status': 'Confirmado',
        'comentario': 'ERRO: NOME E DOC VAZIOS'
    })

    return linhas


def criar_planilha(nome_arquivo="planilha_teste_importacao.xlsx"):
    linhas = gerar_linhas()

    wb = Workbook()
    ws = wb.active
    ws.title = "Importação"

    # ── Estilos ──
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    comment_font = Font(name="Calibri", italic=True, color="888888", size=9)

    # ── Cabeçalho ──
    headers = ["NOME", "CPF_CNPJ", "TELEFONE", "STATUS", "COMENTÁRIO (não enviar)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ── Dados ──
    confirmado_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    tentativa_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    erro_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    for row_idx, linha in enumerate(linhas, 2):
        valores = [linha['nome'], linha['cpf_cnpj'], linha['telefone'], linha['status'], linha['comentario']]
        is_erro = 'ERRO' in linha['comentario']

        for col_idx, valor in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            if is_erro:
                cell.fill = erro_fill

            # Coluna COMENTÁRIO
            if col_idx == 5:
                cell.font = comment_font

            # Coluna STATUS colorida
            if col_idx == 4 and not is_erro:
                if valor == "Confirmado":
                    cell.fill = confirmado_fill
                    cell.font = Font(color="155724")
                elif valor == "Tentativa":
                    cell.fill = tentativa_fill
                    cell.font = Font(color="856404")

    # ── Larguras ──
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 48

    # ── Salva ──
    caminho = os.path.join(os.path.dirname(os.path.abspath(__file__)), nome_arquivo)
    wb.save(caminho)

    print(f"\n✅ Planilha gerada com sucesso: {caminho}")
    print(f"   → {len(linhas)} linhas de dados (+ 1 cabeçalho)")
    print(f"\n📋 Resumo dos cenários de teste:")
    for i, l in enumerate(linhas, 1):
        print(f"   {i:2d}. {l['comentario']:50s} | {l['nome'][:30] or '(vazio)':30s} | {l['status'] or '-'}")


if __name__ == "__main__":
    criar_planilha()
